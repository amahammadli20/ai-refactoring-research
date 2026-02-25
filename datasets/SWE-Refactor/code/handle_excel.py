import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ExcelUtil:
    def __init__(self, file_path=None):
        """
        初始化工具类。
        :param file_path: 文件路径。如果提供路径，则尝试加载文件；否则创建新工作簿。
        """
        self.file_path = file_path
        self.workbook = None

        if file_path:
            try:
                self.workbook = openpyxl.load_workbook(file_path)
                print(f"成功加载文件: {file_path}")
            except FileNotFoundError:
                print(f"文件未找到，创建新工作簿: {file_path}")
                self.workbook = Workbook()
        else:
            self.workbook = Workbook()

    def create_sheet(self, sheet_name):
        """
        创建新工作表。
        :param sheet_name: 工作表名称。
        """
        if sheet_name in self.workbook.sheetnames:
            print(f"工作表 '{sheet_name}' 已存在。")
        else:
            self.workbook.create_sheet(sheet_name)
            print(f"创建工作表: {sheet_name}")

    def write_cell(self, sheet_name, row, column, value):
        """
        向指定单元格写入值。
        :param sheet_name: 工作表名称。
        :param row: 行号（从1开始）。
        :param column: 列号（从1开始）。
        :param value: 写入的值。
        """
        if sheet_name not in self.workbook.sheetnames:
            print(f"工作表 '{sheet_name}' 不存在，无法写入数据。")
            return
        sheet = self.workbook[sheet_name]
        sheet.cell(row=row, column=column, value=value)

    def read_cell(self, sheet_name, row, column):
        """
        读取指定单元格的值。
        :param sheet_name: 工作表名称。
        :param row: 行号（从1开始）。
        :param column: 列号（从1开始）。
        :return: 单元格的值。
        """
        if sheet_name not in self.workbook.sheetnames:
            print(f"工作表 '{sheet_name}' 不存在，无法读取数据。")
            return None
        sheet = self.workbook[sheet_name]
        return sheet.cell(row=row, column=column).value

    def update_cell(self, sheet_name, row, column, value):
        """
        更新指定单元格的值。
        :param sheet_name: 工作表名称。
        :param row: 行号（从1开始）。
        :param column: 列号（从1开始）。
        :param value: 更新的值。
        """
        self.write_cell(sheet_name, row, column, value)

    def save(self, file_path=None):
        """
        保存工作簿。
        :param file_path: 保存的文件路径。如果未提供，则覆盖初始化时的路径。
        """
        path = file_path or self.file_path
        if not path:
            print("未指定保存路径，无法保存文件。")
            return
        self.workbook.save(path)
        print(f"文件已保存: {path}")

    def list_sheets(self):
        """
        列出所有工作表名称。
        :return: 工作表名称列表。
        """
        return self.workbook.sheetnames

    def delete_sheet(self, sheet_name):
        """
        删除指定工作表。
        :param sheet_name: 工作表名称。
        """
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            self.workbook.remove(sheet)
            print(f"已删除工作表: {sheet_name}")
        else:
            print(f"工作表 '{sheet_name}' 不存在，无法删除。")
